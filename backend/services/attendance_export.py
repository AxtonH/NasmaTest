"""Export the team attendance report as .xlsx or .pdf.

This is a faithful port of the standalone Attendance Dashboard's export feature
(`features/exports/{service,excel,pdf}.py`). The `ExportTable` structure, the
openpyxl/reportlab renderers, the styling (dark header band, section shading,
zebra striping), the column shapes, and the filename convention are reproduced
verbatim so a file exported from Nasma matches one exported from the dashboard.

The only Nasma-specific piece is the adapter (`build_daily_export` /
`build_range_export`) that converts Nasma's `MemberRange` data into the same
`ExportTable` row matrix the dashboard builds from its response models. Status
words (Absent / On leave / Holiday) and worked-time formatting match the
dashboard's `exports/service.py` exactly.

No attendance logic here — just reshaping + layout.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from typing import List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

try:  # reuse the formatters + status constants from the report module
    from .attendance_report import (
        MemberRange,
        DayRow,
        STATUS_ABSENT,
        STATUS_ON_LEAVE,
        STATUS_HOLIDAY,
        _fmt_time,
        _fmt_worked_minutes,
        _fmt_weekday_date,
    )
except Exception:  # script-style import
    from attendance_report import (  # type: ignore
        MemberRange,
        DayRow,
        STATUS_ABSENT,
        STATUS_ON_LEAVE,
        STATUS_HOLIDAY,
        _fmt_time,
        _fmt_worked_minutes,
        _fmt_weekday_date,
    )


# --------------------------------------------------------------------------- #
# ExportTable (ported verbatim from the dashboard's exports/service.py)
# --------------------------------------------------------------------------- #

@dataclass(frozen=True)
class ExportTable:
    """A flat, renderer-agnostic view of an attendance report.

    `title` heads the sheet/page. `columns` are the header cells. `rows` is a
    list of string rows aligned to `columns`. `section_rows` holds the indices
    in `rows` that are section headers (per-employee banners in the grouped
    layout) so renderers can style them distinctly; it's empty for the flat
    daily layout. `subtitle` is an optional one-line caption.
    """

    title: str
    subtitle: str
    columns: List[str]
    rows: List[List[str]]
    section_rows: Set[int] = field(default_factory=set)


# --------------------------------------------------------------------------- #
# Status / worked-time formatting (matches the dashboard's exports/service.py)
# --------------------------------------------------------------------------- #

STATUS_ABSENT_WORD = "Absent"
STATUS_ON_LEAVE_WORD = "On leave"
STATUS_HOLIDAY_WORD = "Holiday"


def _worked_or_status(row: DayRow) -> str:
    """One cell combining worked time and excused/absent status (text words).

    Precedence Holiday > On leave > Absent, same as the dashboard. A worked day
    shows the formatted duration.
    """
    if row.status == STATUS_HOLIDAY:
        return STATUS_HOLIDAY_WORD
    if row.status == STATUS_ON_LEAVE:
        return STATUS_ON_LEAVE_WORD
    if row.status == STATUS_ABSENT:
        return STATUS_ABSENT_WORD
    return _fmt_worked_minutes(row.worked_minutes)


# --------------------------------------------------------------------------- #
# Adapter: Nasma MemberRange[] → ExportTable (matches the dashboard layouts)
# --------------------------------------------------------------------------- #

def build_daily_export(ranges: List[MemberRange], *, period_label: str) -> ExportTable:
    """Flat one-row-per-member table for a single day.

    Column shape matches the dashboard's daily Employees export: Emp code, Name,
    Punch in, Punch out, Worked time. Only members scheduled that day appear
    (ranges already exclude unscheduled members, who have no day rows).
    """
    columns = ["Emp code", "Name", "Punch in", "Punch out", "Worked time"]
    rows: List[List[str]] = []
    for mr in ranges:
        if not mr.days:
            continue
        d = mr.days[0]
        rows.append([
            mr.emp_code,
            mr.name,
            _fmt_time(d.punch_in),
            _fmt_time(d.punch_out),
            _worked_or_status(d),
        ])
    return ExportTable(
        title="Attendance · Team",
        subtitle=period_label,
        columns=columns,
        rows=rows,
    )


def build_range_export(ranges: List[MemberRange], *, period_label: str) -> ExportTable:
    """Grouped table for a date range — one section per member.

    A banner row (emp code, name, days-worked + total hours) followed by the
    per-day breakdown. Banner row indices are recorded in `section_rows` so the
    renderers bold/shade them. Mirrors the dashboard's `build_range_export`.
    """
    columns = ["Emp code / Day", "Name", "Punch in", "Punch out", "Worked time"]
    rows: List[List[str]] = []
    section_rows: Set[int] = set()

    for mr in ranges:
        days_label = "1 day" if mr.days_worked == 1 else f"{mr.days_worked} days"
        total = _fmt_worked_minutes(mr.total_worked_minutes)
        section_rows.add(len(rows))
        rows.append([mr.emp_code, mr.name, "", "", f"{days_label} · {total}"])
        for d in mr.days:
            rows.append([
                _fmt_weekday_date(d.day),
                "",
                _fmt_time(d.punch_in),
                _fmt_time(d.punch_out),
                _worked_or_status(d),
            ])

    return ExportTable(
        title="Attendance · Team",
        subtitle=period_label,
        columns=columns,
        rows=rows,
        section_rows=section_rows,
    )


def build_export_table(
    ranges: List[MemberRange],
    start_date: date,
    end_date: date,
) -> ExportTable:
    """Pick the daily or range layout and stamp the period caption."""
    if start_date == end_date:
        return build_daily_export(ranges, period_label=_period_label(start_date, end_date, True))
    return build_range_export(ranges, period_label=_period_label(start_date, end_date, False))


def _period_label(start: date, end: date, single_day: bool) -> str:
    """Human-readable caption, DD-MM-YYYY (matches the dashboard route)."""
    if single_day:
        return start.strftime("%A, %d-%m-%Y")
    return f"{start.strftime('%d-%m-%Y')} to {end.strftime('%d-%m-%Y')}"


def export_filename(start: date, end: date, ext: str) -> str:
    """Descriptive, range-stamped download filename (ISO dates, sortable).

    Single day → prezlab-attendance_2026-06-29.xlsx
    Range      → prezlab-attendance_2026-06-01_to_2026-06-30.pdf
    """
    if start == end:
        stamp = start.isoformat()
    else:
        stamp = f"{start.isoformat()}_to_{end.isoformat()}"
    return f"prezlab-attendance_{stamp}.{ext}"


# --------------------------------------------------------------------------- #
# Excel renderer (ported verbatim from the dashboard's exports/excel.py)
# --------------------------------------------------------------------------- #

_HEADER_FILL = PatternFill("solid", fgColor="1F2430")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="EEF1F5")
_SECTION_FONT = Font(bold=True, color="1F2430")
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(size=10, color="6B7280")
_COLUMN_WIDTHS = [18, 26, 12, 12, 18]


def render_xlsx(table: ExportTable) -> bytes:
    """Build the workbook and return it as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    n_cols = len(table.columns)
    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = table.title
    title_cell.font = _TITLE_FONT

    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = table.subtitle
    subtitle_cell.font = _SUBTITLE_FONT

    header_row_idx = 4  # leave row 3 blank as a spacer
    for col_idx, name in enumerate(table.columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for offset, row_values in enumerate(table.rows):
        excel_row = header_row_idx + 1 + offset
        is_section = offset in table.section_rows
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if is_section:
                cell.fill = _SECTION_FILL
                cell.font = _SECTION_FONT

    _apply_widths(ws, n_cols)
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _apply_widths(ws: Worksheet, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        width = (
            _COLUMN_WIDTHS[col_idx - 1]
            if col_idx - 1 < len(_COLUMN_WIDTHS)
            else 16
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# --------------------------------------------------------------------------- #
# PDF renderer (ported verbatim from the dashboard's exports/pdf.py)
# --------------------------------------------------------------------------- #

_HEADER_BG = colors.HexColor("#1F2430")
_HEADER_FG = colors.white
_SECTION_BG = colors.HexColor("#EEF1F5")
_GRID = colors.HexColor("#E5E7EB")
_ROW_ALT = colors.HexColor("#FAFAFA")
_COL_WIDTHS_MM = [55, 80, 35, 35, 55]


def render_pdf(table: ExportTable) -> bytes:
    """Build the PDF and return it as bytes."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        title=table.title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Title"],
        fontSize=16,
        alignment=0,  # left
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "ExportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#6B7280"),
        spaceAfter=10,
    )

    elements: list = [
        Paragraph(table.title, title_style),
        Paragraph(table.subtitle, subtitle_style),
        Spacer(1, 4),
    ]

    if table.rows:
        elements.append(_build_pdf_table(table))
    else:
        elements.append(Paragraph("No data for this period.", subtitle_style))

    doc.build(elements)
    return buffer.getvalue()


def _build_pdf_table(table: ExportTable) -> Table:
    data = [table.columns, *table.rows]
    col_widths = [w * mm for w in _COL_WIDTHS_MM[: len(table.columns)]]

    style_commands: list = [
        ("BACKGROUND", (0, 0), (-1, 0), _HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), _HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.5, _GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, _GRID),
    ]

    for row_idx in table.section_rows:
        data_row = row_idx + 1  # +1 shifts past the header row
        style_commands.append(("BACKGROUND", (0, data_row), (-1, data_row), _SECTION_BG))
        style_commands.append(("FONTNAME", (0, data_row), (-1, data_row), "Helvetica-Bold"))

    if not table.section_rows:
        for i in range(1, len(data)):
            if i % 2 == 0:
                style_commands.append(("BACKGROUND", (0, i), (-1, i), _ROW_ALT))

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_commands))
    return tbl


# --------------------------------------------------------------------------- #
# Orchestrator (single entry point for the export route)
# --------------------------------------------------------------------------- #

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MEDIA = "application/pdf"


def get_team_attendance_export(
    odoo_service,
    employee_service,
    start_date: date,
    end_date: date,
    fmt: str,
) -> Tuple[bool, Any]:
    """Build the export file. Returns (ok, (bytes, media_type, filename)) or (False, err).

    Reuses the exact same data gathering as the on-screen report so the file
    matches what the manager sees.
    """
    try:
        # Imported lazily to avoid a circular import at module load.
        try:
            from .attendance_report import _gather_member_ranges
        except Exception:
            from attendance_report import _gather_member_ranges  # type: ignore

        ok, ranges, _today = _gather_member_ranges(
            odoo_service, employee_service, start_date, end_date
        )
        if not ok:
            return False, ranges

        table = build_export_table(ranges, start_date, end_date)

        if fmt == "pdf":
            payload = render_pdf(table)
            return True, (payload, _PDF_MEDIA, export_filename(start_date, end_date, "pdf"))
        payload = render_xlsx(table)
        return True, (payload, _XLSX_MEDIA, export_filename(start_date, end_date, "xlsx"))
    except Exception as e:
        return False, f"Error building attendance export: {e}"
