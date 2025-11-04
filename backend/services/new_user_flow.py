from typing import Dict, Any, List, Tuple
from flask import session
import io
from datetime import date
from openpyxl import load_workbook
from .manager_helper import _make_odoo_request


def start_new_user_flow() -> Dict[str, Any]:
    """Initial prompt for creating new Odoo user profiles that opens the upload widget immediately."""
    return {
        'message': ' ',
        'widgets': {
            'new_user_upload': True
        }
    }


def handle_new_user_action(action: str) -> Dict[str, Any]:
    """Handle follow-up action button clicks for the new user flow (placeholder)."""
    action = (action or '').strip().lower()
    if action == 'new_user_manual':
        return {
            'message': (
                "Manual entry flow selected. I'll guide you to add user details step-by-step.\n"
                "(This flow is being set up — for now, please provide: Full name, Email, Job Title, Department.)"
            )
        }
    if action == 'new_user_upload':
        return {
            'message': ' ',
            'widgets': { 'new_user_upload': True }
        }
    return { 'message': "Sorry, I didn't catch that selection. Please choose an option." }


HEADER_TO_FIELD = {
    'Email Address': ('work_email', 'hr.employee'),
    'First and last name as per your passport (in English)': ('name', 'hr.employee'),
    'Date of birth': ('birthday', 'hr.employee'),
    'Personal phone number': ('private_phone', 'hr.employee'),
    'Home Address': ('private_street', 'hr.employee'),
    'Current Address': ('x_studio_work_location_country', 'hr.employee'),
    'Attach a copy of your passport': ('x_studio_passport', 'hr.employee'),
    'Attach your national ID': ('x_studio_national_id_1', 'hr.employee'),
    'Attach your non-criminal certificate': ('x_studio_non_criminal_certificate', 'hr.employee'),
    'What is the position you are hired for?': ('job_id', 'hr.employee'),
    'Bank account No.': ('bank_account_id', 'hr.employee'),
    'Attach your COVID Vaccination Certificate': ('x_studio_covid_vaccination', 'hr.employee'),
    'Emergency Contact Relationship': ('x_studio_contact_1_relation', 'hr.employee'),
    'Marital Status': ('marital', 'hr.employee'),
    'Religion': ('x_studio_religion', 'hr.employee'),
    'Emergency contact first and last name': ('emergency_contact', 'hr.employee'),
    'Emergency contact number': ('emergency_phone', 'hr.employee'),
    'First and last name as per your passport (in Arabic)': ('x_studio_employee_arabic_name', 'hr.employee'),
    'Attach your most updated CV': ('x_studio_cv', 'hr.employee'),
}

# Predefined company options shown in the UI
COMPANY_OPTIONS: List[str] = [
    'Prezlab FZ LLC',
    'Prezlab Advanced Design Company',
    'Prezlab FZ LLC - Regional Office',
    'Prezlab Digital Design Firm L.L.C. - O.P.C',
    'ALOROD AL TAQADAMIAH LEL TASMEM CO',
]

def _normalize_company_label(value: str) -> str:
    return ' '.join(str(value or '').strip().lower().split())

_COMPANY_LOOKUP = { _normalize_company_label(c): c for c in COMPANY_OPTIONS }


def parse_new_user_excel(file_bytes: bytes, odoo_service=None) -> Dict[str, Any]:
    """Parse uploaded Excel, map headers to Odoo fields, return structured rows for confirmation.
    
    Args:
        file_bytes: Excel file content
        odoo_service: Optional OdooService instance for duplicate checking
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return { 'success': False, 'message': 'Excel file is empty' }
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    # Map column indices to field names
    col_to_field = {}
    for idx, h in enumerate(header):
        if h in HEADER_TO_FIELD:
            col_to_field[idx] = HEADER_TO_FIELD[h][0]
    if not col_to_field:
        return { 'success': False, 'message': 'No recognized headers found in the first row' }

    extracted = []
    for r in rows[1:]:
        if not r or all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        rec: Dict[str, Any] = {}
        for idx, val in enumerate(r):
            if idx in col_to_field:
                field = col_to_field[idx]
                if val is None:
                    continue
                # Convert dates to YYYY-MM-DD if needed
                if field == 'birthday' and hasattr(val, 'strftime'):
                    rec[field] = val.strftime('%Y-%m-%d')
                else:
                    rec[field] = str(val).strip()
        if rec:
            extracted.append(rec)

    if not extracted:
        return { 'success': False, 'message': 'No data rows found below header' }

    # Check for duplicate names in Odoo if service is provided
    if odoo_service:
        try:
            odoo_service.ensure_active_session()
        except Exception:
            pass
        for rec in extracted:
            name = rec.get('name', '').strip()
            if name:
                ok, res = _make_odoo_request(
                    odoo_service,
                    'hr.employee',
                    'search_read',
                    {
                        'args': [[('name', '=', name)]],
                        'kwargs': {'fields': ['id', 'name'], 'limit': 1}
                    }
                )
                if ok and isinstance(res, list) and len(res) > 0:
                    rec['_duplicate'] = True
                    rec['_error'] = 'Duplicate name'

    # Save to session for confirmation step
    session['new_user_batch'] = extracted
    return { 'success': True, 'rows': extracted }


def confirmation_message(rows: Any) -> str:
    # Prefer human-friendly header labels when available
    field_to_label = { field: header for header, (field, _model) in HEADER_TO_FIELD.items() }
    # Preferred display order using known headers
    preferred_order = [f for f, _m in [HEADER_TO_FIELD[h] for h in HEADER_TO_FIELD]]
    # Attachment-like fields to compress
    attachment_fields = {
        'x_studio_passport',
        'x_studio_national_id_1',
        'x_studio_non_criminal_certificate',
        'x_studio_covid_vaccination',
        'x_studio_cv',
    }

    def _display_value(key: str, val: Any) -> str:
        if val is None:
            return ''
        s = str(val).strip()
        if not s:
            return ''
        if key in attachment_fields:
            # Show a compact indicator instead of long URL
            return 'Attached'
        if key == 'bank_account_id':
            digits = ''.join(ch for ch in s if ch.isdigit())
            if len(digits) >= 4:
                return f"•••• {digits[-4:]}"
            return s
        return s

    lines = []
    
    # Separate duplicates from valid users
    duplicates = [rec for rec in rows if rec.get('_duplicate')]
    valid_users = [rec for rec in rows if not rec.get('_duplicate')]
    
    # Show duplicates first in red at the top
    if duplicates:
        lines.append('<span style="color: red; font-weight: bold;">⚠️ INVALID USERS - DUPLICATE NAMES:</span>')
        for rec in duplicates:
            name = rec.get('name', 'Unknown')
            error = rec.get('_error', 'Duplicate name')
            lines.append(f'<span style="color: red;">❌ {name}: {error}</span>')
        lines.append('')  # Empty line for spacing
    
    # Show valid users
    if valid_users:
        lines.append("Please confirm the following new user records:")
        for i, rec in enumerate(valid_users, start=1):
            lines.append(f"\n{i})")
            # Use preferred order first, then any remaining keys
            keys_in_order = []
            seen = set()
            for k in preferred_order:
                if k in rec and not k.startswith('_'):  # Skip internal flags
                    keys_in_order.append(k)
                    seen.add(k)
            for k in rec.keys():
                if k not in seen and not k.startswith('_'):  # Skip internal flags
                    keys_in_order.append(k)

            for k in keys_in_order:
                v = _display_value(k, rec.get(k))
                if v == '':
                    continue
                label = field_to_label.get(k, k)
                lines.append(f"- **{label}**: {v}")
    
    if duplicates and not valid_users:
        lines.append("\n<span style=\"color: red;\">All users have duplicate names. Please fix the issues and try again.</span>")
    elif duplicates and valid_users:
        lines.append("\n<span style=\"color: orange;\">⚠️ Only valid users will be created if you confirm. Flagged users will be skipped.</span>")
        lines.append("\nClick Confirm to create valid users or Cancel to abort.")
    elif valid_users:
        lines.append("\nClick Confirm to create users or Cancel to abort.")
    
    return "\n".join(lines)


def confirm_new_user_batch() -> Dict[str, Any]:
    batch = session.get('new_user_batch') or []
    if not batch:
        return { 'message': 'No pending new users to create.' }
    # Placeholder for Odoo creation; integrate later
    session.pop('new_user_batch', None)
    return { 'message': f"Successfully queued {len(batch)} new user(s) for creation.", 'success': True }


 


def _normalize_marital(value: str) -> str:
    if not value:
        return ''
    v = str(value).strip().lower()
    aliases = {
        'single': 'single',
        'married': 'married',
        'cohabitant': 'cohabitant',
        'widower': 'widower', 'widowed': 'widower',
        'divorced': 'divorced'
    }
    return aliases.get(v, v)


def _resolve_job_id(odoo_service, name: str) -> int:
    if not name:
        return 0
    ok, res = _make_odoo_request(
        odoo_service,
        'hr.job',
        'search_read',
        {
            'args': [[('name', 'ilike', name)]],
            'kwargs': {'fields': ['id', 'name'], 'limit': 1}
        }
    )
    if ok and isinstance(res, list) and res:
        rid = res[0].get('id')
        return int(rid) if isinstance(rid, int) else 0
    return 0


def _resolve_company_id(odoo_service, name: str) -> int:
    if not name:
        return 0
    ok, res = _make_odoo_request(
        odoo_service,
        'res.company',
        'search_read',
        {
            'args': [[('name', '=', name)]],
            'kwargs': {'fields': ['id', 'name'], 'limit': 1}
        }
    )
    if ok and isinstance(res, list) and res:
        rid = res[0].get('id')
        return int(rid) if isinstance(rid, int) else 0
    # Fallback to ilike search
    ok2, res2 = _make_odoo_request(
        odoo_service,
        'res.company',
        'search_read',
        {
            'args': [[('name', 'ilike', name)]],
            'kwargs': {'fields': ['id', 'name'], 'limit': 1}
        }
    )
    if ok2 and isinstance(res2, list) and res2:
        rid = res2[0].get('id')
        return int(rid) if isinstance(rid, int) else 0
    return 0


def assign_company_to_record(index: int, company_label: str, odoo_service) -> Dict[str, Any]:
    """Assign company to a pending new user record stored in session.

    Args:
        index: zero-based index into session['new_user_batch']
        company_label: company display name as shown in the dropdown
        odoo_service: active Odoo service to resolve company id
    Returns:
        dict with success flag and updated rows
    """
    batch = session.get('new_user_batch') or []
    if not isinstance(index, int) or index < 0 or index >= len(batch):
        return { 'success': False, 'message': 'Invalid user index' }

    canonical = _COMPANY_LOOKUP.get(_normalize_company_label(company_label))
    if not canonical:
        return { 'success': False, 'message': 'Unknown company selected' }

    ok_session, msg = odoo_service.ensure_active_session()
    if not ok_session:
        return { 'success': False, 'message': f'Odoo session error: {msg}' }

    cid = _resolve_company_id(odoo_service, canonical)
    if not cid:
        return { 'success': False, 'message': f"Company '{canonical}' not found in Odoo" }

    # Persist selection in session for later creation
    batch[index]['company_id'] = cid
    batch[index]['company_name'] = canonical
    session['new_user_batch'] = batch
    return { 'success': True, 'rows': batch }


def create_employees_batch(odoo_service) -> Dict[str, Any]:
    """Create hr.employee records based on parsed batch in session.

    Note: Attachments and bank accounts are not created yet; those fields are skipped for now.
    Records flagged as duplicates (_duplicate=True) will be skipped.
    """
    batch = session.get('new_user_batch') or []
    if not batch:
        return { 'success': False, 'message': 'No pending new users to create.' }

    # Ensure session is active once at the start
    ok_session, session_msg = odoo_service.ensure_active_session()
    if not ok_session:
        return { 'success': False, 'message': f'Session error: {session_msg}' }

    # Separate valid users from flagged duplicates
    valid_batch = [rec for rec in batch if not rec.get('_duplicate')]
    skipped_duplicates = [rec for rec in batch if rec.get('_duplicate')]

    # If all users are duplicates, return error
    if not valid_batch and skipped_duplicates:
        session.pop('new_user_batch', None)
        dup_names = [rec.get('name', 'Unknown') for rec in skipped_duplicates]
        return {
            'success': False,
            'message': f"❌ Cannot create users. All names are duplicates: {', '.join(dup_names)}\n\nPlease fix the duplicate names and try again."
        }

    # Validate company assignment for all valid users
    missing_company = [rec for rec in valid_batch if not rec.get('company_id')]
    if missing_company:
        names = [rec.get('name', 'Unknown') for rec in missing_company]
        return {
            'success': False,
            'message': (
                "❌ Cannot create users: company is required for all users.\n"
                f"Missing company for: {', '.join(names)}\n\n"
                "Please assign a company to each user (use 'Assign company')."
            )
        }

    created: List[Dict[str, Any]] = []
    failed: List[Dict[str, Any]] = []

    for rec in valid_batch:
        try:
            vals: Dict[str, Any] = {}
            # Direct text fields
            # Note: x_studio_work_location_country is handled separately as a many2one field below
            for field in [
                'name','work_email','birthday','private_phone','private_street',
                'x_studio_contact_1_relation','marital','x_studio_religion',
                'emergency_contact','emergency_phone','x_studio_employee_arabic_name'
            ]:
                if rec.get(field):
                    if field == 'marital':
                        vals[field] = _normalize_marital(rec.get(field))
                    else:
                        vals[field] = rec.get(field)

            # job_id by fuzzy name search
            if rec.get('job_id'):
                jid = _resolve_job_id(odoo_service, rec.get('job_id'))
                if jid:
                    vals['job_id'] = jid

            # Required: company_id
            if rec.get('company_id'):
                vals['company_id'] = int(rec['company_id'])

            # Many2one: x_studio_work_location_country -> res.country
            # Handle cases where value might be "City, Country" format
            if rec.get('x_studio_work_location_country'):
                country_value = str(rec.get('x_studio_work_location_country')).strip()
                # If value contains comma, try to extract country name (last part after comma)
                if ',' in country_value:
                    parts = [p.strip() for p in country_value.split(',')]
                    country_name = parts[-1] if parts else country_value
                else:
                    country_name = country_value
                
                ok_c, res_c = _make_odoo_request(
                    odoo_service,
                    'res.country',
                    'search_read',
                    {
                        'args': [[('name', 'ilike', country_name)]],
                        'kwargs': {'fields': ['id', 'name'], 'limit': 1}
                    }
                )
                if ok_c and isinstance(res_c, list) and res_c:
                    cid = res_c[0].get('id')
                    if isinstance(cid, int):
                        vals['x_studio_work_location_country'] = cid
                # If not found, drop the string to avoid type error (field won't be set)

            # TODO: bank_account_id and attachments require additional flows; skip for now

            ok, result = _make_odoo_request(
                odoo_service,
                'hr.employee',
                'create',
                {
                    'args': [vals],
                    'kwargs': {}
                }
            )
            if ok and isinstance(result, int):
                created.append({'id': result, 'name': vals.get('name', '')})
            else:
                failed.append({'name': vals.get('name', ''), 'error': str(result)})
        except Exception as e:
            failed.append({'name': rec.get('name', ''), 'error': str(e)})

    # Clear batch after attempt
    session.pop('new_user_batch', None)

    # Generate friendly confirmation messages
    messages = []
    attachments: List[Dict[str, Any]] = []
    
    # Show skipped duplicates first
    if skipped_duplicates:
        messages.append("⚠️ **Skipped duplicate names:**")
        for dup in skipped_duplicates:
            name = dup.get('name', 'Unknown')
            messages.append(f"   ❌ {name} (already exists in Odoo)")
        messages.append("")  # Empty line for spacing
    
    # Show successfully created users
    if created:
        for idx, c in enumerate(created):
            name = c.get('name', 'Unknown')
            messages.append(f"🎉 {name} has been added to the Prezlab family!")
            try:
                # If company is one of the specified companies, generate service agreement using company-specific template
                rec = valid_batch[idx] if idx < len(valid_batch) else {}
                comp_name = rec.get('company_name') or ''
                if comp_name in {
                    'Prezlab FZ LLC - Regional Office',
                    'ALOROD AL TAQADAMIAH LEL TASMEM CO',
                    'Prezlab Advanced Design Company',
                    'Prezlab FZ LLC',
                    'Prezlab Digital Design Firm L.L.C. - O.P.C'
                }:
                    from .employee_service import EmployeeService
                    from .document_service import DocumentService
                    # Build minimal services using shared odoo_service
                    emp_service = EmployeeService(odoo_service)
                    doc_service = DocumentService(odoo_service, emp_service)
                    ok_doc, doc_meta = doc_service.generate_service_agreement(
                        name,
                        private_street=rec.get('private_street') or rec.get('Private Street') or '',
                        company_name=comp_name
                    )
                    if ok_doc and isinstance(doc_meta, dict):
                        attachments.append({
                            'file_url': doc_meta.get('file_url'),
                            'file_name': doc_meta.get('file_name')
                        })
            except Exception:
                pass
    
    # Show failed users
    if failed:
        for f in failed:
            name = f.get('name', 'Unknown')
            error = f.get('error', 'Unknown error')
            messages.append(f"❌ Failed to add {name}: {error}")
    
    hardware_options: List[Dict[str, Any]] = []
    if created:
        for rec in created:
            full_name = rec.get('name', '').strip()
            first_name = full_name.split()[0] if full_name else 'New user'
            hardware_options.append({
                'employee_id': rec.get('id'),
                'name': full_name,
                'first_name': first_name
            })
        session['new_user_recent_employees'] = hardware_options

    return { 
        'success': True, 
        'message': '\n'.join(messages) if messages else 'No employees processed.', 
        'created': created, 
        'failed': failed,
        'skipped': skipped_duplicates,
        'attachments': attachments,
        'hardware_options': hardware_options
    }


def list_available_hardware(odoo_service, limit: int = 200) -> List[Dict[str, Any]]:
    """Return available maintenance equipment (not assigned to any employee)."""
    args = [[('employee_id', '=', False)]]
    kwargs = {'fields': ['id', 'name'], 'limit': limit}
    ok, res = _make_odoo_request(
        odoo_service,
        'maintenance.equipment',
        'search_read',
        {'args': args, 'kwargs': kwargs}
    )
    if not ok or not isinstance(res, list):
        return []
    hardware_list: List[Dict[str, Any]] = []
    for item in res:
        hid = item.get('id')
        name = (item.get('name') or '').strip()
        if hid and name:
            hardware_list.append({'id': hid, 'name': name})
    return hardware_list


def assign_hardware_to_employee(odoo_service, hardware_id: int, employee_id: int) -> Tuple[bool, str]:
    """Assign maintenance equipment to the given employee with today's date."""
    try:
        hardware_id = int(hardware_id)
        employee_id = int(employee_id)
    except Exception:
        return False, 'Invalid identifiers for hardware assignment.'

    ok_session, session_msg = odoo_service.ensure_active_session()
    if not ok_session:
        return False, f'Odoo session error: {session_msg}'

    payload = {
        'equipment_assign_to': 'employee',
        'employee_id': employee_id,
        'assign_date': date.today().isoformat()
    }
    ok, res = _make_odoo_request(
        odoo_service,
        'maintenance.equipment',
        'write',
        {'args': [[hardware_id], payload], 'kwargs': {}}
    )
    if not ok:
        return False, str(res)
    return True, ''
