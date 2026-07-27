"""Render an `AttendanceReport` to an .xlsx workbook (openpyxl).

Nasma port of the Attendance Dashboard's `features/exports/excel.py`,
reproduced verbatim (only the import style differs) so a file exported from
Nasma matches one exported from the dashboard.

The only place that knows the Excel format. The Summary sheet follows the
Prezlab brand template: a deep-teal header band, four hero KEY METRICS
tiles (the first inverted), side-by-side arrival/departure bucket tables
each with a native bar chart, the per-employee overview with hyperlinks,
and four TOP RANKINGS podiums. Each employee sheet follows the same system:
header band with a back-link to the summary, PROFILE, a SUMMARY tile grid
(three tiles per band, the attendance tile highlighted), the DAILY LOG with
a colored status pill per day, and a STATUS KEY legend. Both sheet types sit
on a B:I grid with a narrow column A gutter and gridlines off, so they read
as designed pages rather than spreadsheets.

No attendance logic here — just layout.
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

try:  # package-style import (matches app.py's primary import block)
    from .attendance_export_report import (
        DASH,
        STATUS_ABSENT,
        STATUS_DAY_OFF,
        STATUS_HOLIDAY,
        STATUS_LATE,
        STATUS_ON_LEAVE,
        STATUS_PRESENT,
        STATUS_UPCOMING,
        STATUS_WEEKEND,
        STATUS_WFH,
        AttendanceReport,
        EmployeeSection,
    )
except Exception:  # script-style import (running from backend/)
    from attendance_export_report import (  # type: ignore
        DASH,
        STATUS_ABSENT,
        STATUS_DAY_OFF,
        STATUS_HOLIDAY,
        STATUS_LATE,
        STATUS_ON_LEAVE,
        STATUS_PRESENT,
        STATUS_UPCOMING,
        STATUS_WEEKEND,
        STATUS_WFH,
        AttendanceReport,
        EmployeeSection,
    )

# ---------- Visual system (Prezlab brand template) ----------

_INK = "002528"  # Prezlab deep teal — title band, table headers, hero tile
_MINT = "B6EDF3"  # Prezlab mint — hero tiles, band-on-ink label text
_MUTED = "62848C"  # slate — secondary labels, captions
_ALT_BG = "F4F6F6"  # section bands + alternating data rows
_GRID = "D5DEE0"  # thin border color
_ACCENT = _INK  # chart bars

_INK_FILL = PatternFill("solid", fgColor=_INK)
_MINT_FILL = PatternFill("solid", fgColor=_MINT)
_ALT_FILL = PatternFill("solid", fgColor=_ALT_BG)
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

_BAND_FILL = _INK_FILL
_BAND_FONT = Font(bold=True, size=12, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
_TABLE_HEADER_FILL = _INK_FILL
_TABLE_HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
_LABEL_FILL = _MINT_FILL
_LABEL_FONT = Font(bold=True, size=10, color=_INK)
_VALUE_FONT = Font(size=10)
_VALUE_BOLD_FONT = Font(bold=True, size=11)
_DETAIL_FONT = Font(size=9, color=_MUTED)
_BODY_FONT = Font(size=10, color=_INK)
_MUTED_FONT = Font(size=10, color="9CA3AF")
_LINK_FONT = Font(size=10, color=_INK, underline="single", bold=True)

# Summary-sheet specific type scale, matching the brand template.
_BRAND_FONT = Font(bold=True, size=12, color=_MINT)
_HERO_TITLE_FONT = Font(bold=True, size=26, color="FFFFFF")
_HERO_SUB_FONT = Font(size=12, color=_MINT)
_HERO_META_LABEL_FONT = Font(size=9, color=_MINT)
_HERO_META_VALUE_FONT = Font(bold=True, size=11, color="FFFFFF")
_SECTION_FONT = Font(bold=True, size=11, color=_INK)
_SECTION_NOTE_FONT = Font(size=8, color=_MUTED)
_TILE_LABEL_ON_INK = Font(size=9, color=_MINT)
_TILE_LABEL_ON_MINT = Font(size=9, color=_MUTED)
_TILE_VALUE_ON_INK = Font(bold=True, size=30, color="FFFFFF")
_TILE_VALUE_ON_MINT = Font(bold=True, size=30, color=_INK)
_TILE_DETAIL_ON_INK = Font(size=8, color=_MINT)
_TILE_DETAIL_ON_MINT = Font(size=8, color=_MUTED)
_MINI_HEADER_FONT = Font(bold=True, size=9, color="FFFFFF")
_RANK_TITLE_FONT = Font(bold=True, size=10, color=_INK)
_RANK_HEADER_FONT = Font(bold=True, size=8, color=_MUTED)
_RANK_NAME_FONT = Font(size=9, color=_INK)
_RANK_EMPTY_FONT = Font(size=9, color=_MUTED)
_RANK_VALUE_FONT = Font(bold=True, size=9, color=_INK)
_NAME_FONT = Font(bold=True, size=10, color=_INK)
_DEPT_FONT = Font(size=10, color=_MUTED)
_PROFILE_LABEL_FONT = Font(size=9, color=_MUTED)
_PROFILE_VALUE_FONT = Font(bold=True, size=10, color=_INK)
_TILE_CAPTION_FONT = Font(size=8, color=_MUTED)
_EMP_TILE_VALUE_FONT = Font(bold=True, size=15, color=_INK)
_FOOTER_FONT = Font(size=8, color=_MUTED)
_FOOTER_BRAND_FONT = Font(bold=True, size=9, color=_MUTED)

_THIN = Side(style="thin", color=_GRID)
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SECTION_RULE = Border(bottom=Side(style="medium", color=_INK))
# The template rules its tables and tiles in the slate secondary color.
_RULE_SIDE = Side(style="thin", color=_MUTED)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

_BAND_HEIGHT = 24
_TITLE_HEIGHT = 30
_ROW_HEIGHT = 18

# Status pill colors, from the brand template's STATUS KEY legend. Every
# status is a filled pill so the daily log scans by color: (fill, font-color).
_STATUS_STYLES: Dict[str, Tuple[str, str]] = {
    STATUS_PRESENT: (_MINT, _INK),
    STATUS_LATE: ("FFE2B8", "7A4A00"),
    STATUS_WFH: ("DBF5F9", _INK),
    STATUS_ON_LEAVE: ("EFE8DE", _MUTED),
    STATUS_HOLIDAY: ("E1E7E9", _MUTED),
    STATUS_ABSENT: ("CDD6D9", _INK),
    STATUS_WEEKEND: (_ALT_BG, _MUTED),
    STATUS_DAY_OFF: (_ALT_BG, _MUTED),
    STATUS_UPCOMING: (_ALT_BG, "C4C4C8"),
}
# The legend row on each employee sheet, in the template's order. Late is
# included (the template's sample had no late days) so every pill that can
# appear in the daily log is explained by the key.
_STATUS_KEY = [
    STATUS_PRESENT,
    STATUS_LATE,
    STATUS_WFH,
    STATUS_ON_LEAVE,
    STATUS_HOLIDAY,
    STATUS_ABSENT,
    STATUS_WEEKEND,
]
# Rows whose non-status cells read as secondary — nothing happened that day.
_MUTED_STATUSES = (STATUS_WFH, STATUS_WEEKEND, STATUS_DAY_OFF, STATUS_UPCOMING)

_SUMMARY_SHEET = "Summary"
# Excel sheet-name hard limits: 31 chars, none of []:*?/\ .
_SHEET_NAME_BAD = str.maketrans("", "", "[]:*?/\\")
_SHEET_NAME_MAX = 28  # leave room for a " (n)" dedupe suffix


def render_xlsx(report: AttendanceReport) -> bytes:
    """Build the workbook and return it as bytes."""
    wb = Workbook()
    summary = wb.active
    summary.title = _SUMMARY_SHEET

    sheet_names = _sheet_names(report.sections)
    _render_summary(summary, report, sheet_names)
    for section, sheet_name in zip(report.sections, sheet_names):
        ws = wb.create_sheet(title=sheet_name)
        _render_employee_sheet(ws, section, report.period_label)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _sheet_names(sections: List[EmployeeSection]) -> List[str]:
    """Legal, unique sheet name per employee (Excel: ≤31 chars, no []:*?/\\)."""
    seen: Dict[str, int] = {}
    names: List[str] = []
    for s in sections:
        base = (s.name.translate(_SHEET_NAME_BAD).strip() or s.emp_code)[
            :_SHEET_NAME_MAX
        ].strip()
        count = seen.get(base.lower(), 0)
        seen[base.lower()] = count + 1
        names.append(base if count == 0 else f"{base} ({count + 1})")
    return names


# ---------- Shared building blocks ----------


def _band(ws: Worksheet, row: int, n_cols: int, text: str, *, title: bool = False) -> None:
    """A full-width dark band — the sheet title or a section header."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _BAND_FILL
    cell.font = _TITLE_FONT if title else _BAND_FONT
    cell.alignment = _LEFT
    # Fill the merged range so the band has no white seams.
    for col in range(2, n_cols + 1):
        ws.cell(row=row, column=col).fill = _BAND_FILL
    ws.row_dimensions[row].height = _TITLE_HEIGHT if title else _BAND_HEIGHT


def _table_header(ws: Worksheet, row: int, columns: List[str]) -> None:
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = _TABLE_HEADER_FILL
        cell.font = _TABLE_HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BOX
    ws.row_dimensions[row].height = _ROW_HEIGHT + 2


def _set_widths(ws: Worksheet, widths: List[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


# ---------- Summary sheet ----------

# The brand template lays the summary out on a B:I content grid with a
# narrow column A gutter. These are 1-based sheet columns.
#
# Deviation from the template's widths: the arrival block (B:D) and the
# departure block (F:I) both sum to 50 units, so the side-by-side bucket
# tables — and the charts cell-anchored to them — render at the same width
# and read as a matched pair (verified by measuring both spans in Excel via
# COM: 270pt each). The template had 50 vs 46, a visible ~25px mismatch.
_SUM_FIRST = 2  # B
_SUM_LAST = 9  # I
_SUM_WIDTHS = [2.44, 24, 15, 11, 11, 12, 12, 13, 13]

# The four hero tiles sit at these column pairs; the first is inverted
# (ink fill) so the eye lands on the attendance rate first.
_TILE_SPANS = [(2, 3), (4, 5), (6, 7), (8, 9)]


def _paint(ws: Worksheet, row: int, c0: int, c1: int, fill: PatternFill) -> None:
    for col in range(c0, c1 + 1):
        ws.cell(row=row, column=col).fill = fill


def _spacer(ws: Worksheet, row: int, *, height: float) -> None:
    """A short tinted row — breathing space inside a continuous panel."""
    _paint(ws, row, _SUM_FIRST, _SUM_LAST, _ALT_FILL)
    ws.row_dimensions[row].height = height


def _outline(
    ws: Worksheet, r0: int, c0: int, r1: int, c1: int, *, inner_v: bool = False
) -> None:
    """Rule a rectangular block: outer box, optionally vertical inner lines.

    Edges are merged into each cell's existing border rather than assigned,
    so a corner cell keeps both of the edges it sits on. Each inner rule is
    drawn once, as the right edge of the left-hand cell — the same single-
    sided convention the template uses (drawing both sides would render
    identically but double the border records).
    """
    for row in range(r0, r1 + 1):
        for col in range(c0, c1 + 1):
            cell = ws.cell(row=row, column=col)
            current = cell.border
            cell.border = Border(
                left=_RULE_SIDE if col == c0 else current.left,
                right=_RULE_SIDE if (col == c1 or inner_v) else current.right,
                top=_RULE_SIDE if row == r0 else current.top,
                bottom=_RULE_SIDE if row == r1 else current.bottom,
            )


def _section_header(ws: Worksheet, row: int, title: str, note: str = "") -> None:
    """A light band with a rule under it — the template's section divider."""
    _paint(ws, row, _SUM_FIRST, _SUM_LAST, _ALT_FILL)
    for col in range(_SUM_FIRST, _SUM_LAST + 1):
        ws.cell(row=row, column=col).border = _SECTION_RULE
    cell = ws.cell(row=row, column=_SUM_FIRST, value=title)
    cell.font = _SECTION_FONT
    cell.alignment = _LEFT
    if note:
        note_cell = ws.cell(row=row, column=_SUM_LAST, value=note)
        note_cell.font = _SECTION_NOTE_FONT
        note_cell.alignment = _RIGHT


def _render_summary(
    ws: Worksheet, report: AttendanceReport, sheet_names: List[str]
) -> None:
    ws.sheet_view.showGridLines = False
    for idx, width in enumerate(_SUM_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _summary_header(ws, report)
    row = _summary_key_metrics(ws, report, start_row=8)
    row = _summary_patterns(ws, report, start_row=row + 1)
    row = _summary_employees(ws, report, sheet_names, start_row=row + 1)
    # The rankings panel and the footer close the page as one tinted block,
    # so the spacer rows between them are filled rather than left white.
    _spacer(ws, row + 1, height=10.05)
    row = _summary_rankings(ws, report, start_row=row + 2)
    _spacer(ws, row + 1, height=10.05)
    _summary_footer(ws, row + 2)


def _summary_header(ws: Worksheet, report: AttendanceReport) -> None:
    """Rows 2–5: the dark brand band with title, subtitle and period meta."""
    for row in range(2, 6):
        _paint(ws, row, _SUM_FIRST, _SUM_LAST, _INK_FILL)
    ws.row_dimensions[2].height = 19.95
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 18

    brand = ws.cell(row=2, column=_SUM_FIRST, value="prezlab:")
    brand.font = _BRAND_FONT
    brand.alignment = _LEFT

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    title = ws.cell(row=3, column=2, value=report.title)
    title.font = _HERO_TITLE_FONT
    title.alignment = _LEFT

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    subtitle = ws.cell(row=4, column=2, value=report.subtitle)
    subtitle.font = _HERO_SUB_FONT
    subtitle.alignment = _LEFT

    meta = [
        ("Period", report.period_label),
        ("Generated", report.generated_on),
        ("Employees", str(report.employee_count)),
    ]
    for offset, (label, value) in enumerate(meta):
        row = 3 + offset
        label_cell = ws.cell(row=row, column=7, value=label)
        label_cell.font = _HERO_META_LABEL_FONT
        label_cell.alignment = _LEFT
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        value_cell = ws.cell(row=row, column=8, value=value)
        value_cell.font = _HERO_META_VALUE_FONT
        value_cell.alignment = _RIGHT


def _summary_key_metrics(
    ws: Worksheet, report: AttendanceReport, *, start_row: int
) -> int:
    """The four hero tiles: label row, big value row, caption row."""
    _section_header(ws, start_row, "KEY METRICS")
    ws.row_dimensions[start_row + 1].height = 6

    label_row = start_row + 2
    value_row = label_row + 1
    detail_row = value_row + 2  # value spans two rows
    ws.row_dimensions[label_row].height = 16.05
    ws.row_dimensions[value_row].height = 19.95
    ws.row_dimensions[value_row + 1].height = 18
    ws.row_dimensions[detail_row].height = 13.95
    ws.row_dimensions[detail_row + 1].height = 10.05

    for i, (metric, (c0, c1)) in enumerate(zip(report.key_metrics, _TILE_SPANS)):
        inverted = i == 0
        fill = _INK_FILL if inverted else _MINT_FILL
        for row in (label_row, value_row, value_row + 1, detail_row):
            _paint(ws, row, c0, c1, fill)

        ws.merge_cells(
            start_row=label_row, start_column=c0, end_row=label_row, end_column=c1
        )
        label = ws.cell(row=label_row, column=c0, value=metric.label)
        label.font = _TILE_LABEL_ON_INK if inverted else _TILE_LABEL_ON_MINT
        label.alignment = _LEFT

        ws.merge_cells(
            start_row=value_row, start_column=c0, end_row=value_row + 1, end_column=c1
        )
        value = ws.cell(row=value_row, column=c0, value=metric.value)
        value.font = _TILE_VALUE_ON_INK if inverted else _TILE_VALUE_ON_MINT
        value.alignment = _LEFT

        ws.merge_cells(
            start_row=detail_row, start_column=c0, end_row=detail_row, end_column=c1
        )
        detail = ws.cell(row=detail_row, column=c0, value=metric.detail)
        detail.font = _TILE_DETAIL_ON_INK if inverted else _TILE_DETAIL_ON_MINT
        detail.alignment = _LEFT

    return detail_row + 1


def _bucket_table(
    ws: Worksheet,
    *,
    row: int,
    label_col: int,
    label_end: int,
    count_col: int,
    count_end: int,
    heading: str,
    bins: List,
) -> int:
    """One side of the arrival/departure pattern: header + banded rows.

    Both the label and count cells span a column range so the two tables
    line up on the template's B:I grid despite their differing widths.
    """

    def cell(at_row: int, c0: int, c1: int, value, font, align) -> None:
        if c1 > c0:
            ws.merge_cells(
                start_row=at_row, start_column=c0, end_row=at_row, end_column=c1
            )
        target = ws.cell(row=at_row, column=c0, value=value)
        target.font = font
        target.alignment = align

    header_row = row
    _paint(ws, row, label_col, count_end, _INK_FILL)
    cell(row, label_col, label_end, heading, _MINI_HEADER_FONT, _LEFT)
    cell(row, count_col, count_end, "Count", _MINI_HEADER_FONT, _CENTER)

    for i, bin_ in enumerate(bins):
        row += 1
        ws.row_dimensions[row].height = 13.95
        _paint(ws, row, label_col, count_end, _ALT_FILL if i % 2 else _WHITE_FILL)
        cell(row, label_col, label_end, bin_.label, _BODY_FONT, _LEFT)
        cell(row, count_col, count_end, bin_.count, _BODY_FONT, _CENTER)

    # Outer box plus a rule between the label and count columns, drawn as
    # the right edge of the label's last column.
    _outline(ws, header_row, label_col, row, count_end)
    for at_row in range(header_row, row + 1):
        current = ws.cell(row=at_row, column=label_end).border
        ws.cell(row=at_row, column=label_end).border = Border(
            left=current.left,
            right=_RULE_SIDE,
            top=current.top,
            bottom=current.bottom,
        )
    return row


# Rows of the 13.95pt chart band a chart spans — ≈6.4cm, the closest row
# multiple to the template's 6.5cm chart height.
_CHART_ROW_SPAN = 13


def _pattern_chart(
    ws: Worksheet, *, title: str, first_row: int, last_row: int, label_col: int,
    value_col: int, anchor_row: int, span: Tuple[int, int],
) -> None:
    """A native Excel bar chart over one bucket table's rows.

    `span` is the (first, last) column of the chart's own bucket table. The
    chart is anchored cell-to-cell across that span, so Excel pins both of
    its edges to the same column boundaries the tables use and computes the
    width from live column geometry at render time. A fixed extent (what the
    template used) can never do this reliably: the on-screen pixel width of
    a column depends on the viewer's DPI scaling and default font metrics,
    so any hardcoded width is misaligned for some viewer.
    """
    bar = BarChart()
    bar.type = "col"
    bar.title = title
    bar.legend = None
    bar.gapWidth = 60
    bar.y_axis.majorGridlines = None
    # openpyxl leaves `delete` unset on a new chart, which Excel reads as
    # "hide the axis" — the bars would float with no time labels or counts.
    # Both axes are drawn with their labels beside them, as the template has.
    for axis in (bar.x_axis, bar.y_axis):
        axis.delete = False
        axis.tickLblPos = "nextTo"
        axis.numFmt = "General"
    data = Reference(ws, min_col=value_col, min_row=first_row, max_row=last_row)
    cats = Reference(ws, min_col=label_col, min_row=first_row, max_row=last_row)
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = _ACCENT
    bar.series[0].graphicalProperties.line.noFill = True
    # AnchorMarker is 0-based; `to` marks the boundary *after* the span's
    # last column, i.e. exactly the table's right edge.
    first_col, last_col = span
    row0 = anchor_row - 1
    bar.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=first_col - 1, row=row0),
        to=AnchorMarker(col=last_col, row=row0 + _CHART_ROW_SPAN),
    )
    ws._charts.append(bar)


def _summary_patterns(
    ws: Worksheet, report: AttendanceReport, *, start_row: int
) -> int:
    """Arrival and departure distributions, side by side, each with a chart."""
    _section_header(ws, start_row, "ARRIVAL & DEPARTURE PATTERN")
    ws.row_dimensions[start_row + 1].height = 6

    header_row = start_row + 2
    # Arrival occupies B | C:D, departure F:G | H:I — the template's grid.
    arrival_end = _bucket_table(
        ws,
        row=header_row,
        label_col=2,
        label_end=2,
        count_col=3,
        count_end=4,
        heading="Arrival",
        bins=report.arrival_buckets,
    )
    departure_end = _bucket_table(
        ws,
        row=header_row,
        label_col=6,
        label_end=7,
        count_col=8,
        count_end=9,
        heading="Departure",
        bins=report.departure_buckets,
    )
    last_row = max(arrival_end, departure_end)

    chart_row = last_row + 2
    ws.row_dimensions[last_row + 1].height = 6
    if report.arrival_buckets:
        _pattern_chart(
            ws,
            title="Arrivals",
            first_row=header_row + 1,
            last_row=arrival_end,
            label_col=2,
            value_col=3,
            anchor_row=chart_row,
            span=(2, 4),  # B:D — the arrival table's columns
        )
    if report.departure_buckets:
        _pattern_chart(
            ws,
            title="Departures",
            first_row=header_row + 1,
            last_row=departure_end,
            label_col=6,
            value_col=8,
            anchor_row=chart_row,
            span=(6, 9),  # F:I — the departure table's columns
        )
    # The chart band: the anchors span the first _CHART_ROW_SPAN of these
    # rows; the remaining rows are the template's breathing space before the
    # next section (keeps EMPLOYEES on the template's row).
    for row in range(chart_row, chart_row + 15):
        ws.row_dimensions[row].height = 13.95
    return chart_row + 15


def _summary_employees(
    ws: Worksheet,
    report: AttendanceReport,
    sheet_names: List[str],
    *,
    start_row: int,
) -> int:
    """The per-employee overview table, each name linking to its own sheet."""
    _section_header(
        ws, start_row, "EMPLOYEES", note="Select a name to open detail"
    )
    row = start_row + 1
    header_row = row
    for i, name in enumerate(report.overview_columns):
        cell = ws.cell(row=row, column=_SUM_FIRST + i, value=name)
        cell.fill = _INK_FILL
        cell.font = _MINI_HEADER_FONT
        cell.alignment = _LEFT if i <= 1 else _CENTER
    ws.row_dimensions[row].height = _ROW_HEIGHT

    for i, (values, sheet_name) in enumerate(
        zip(report.overview_rows, sheet_names)
    ):
        row += 1
        fill = _ALT_FILL if i % 2 == 1 else _WHITE_FILL
        for j, value in enumerate(values):
            cell = ws.cell(row=row, column=_SUM_FIRST + j, value=value)
            cell.fill = fill
            if j == 0:
                cell.font = _NAME_FONT
            elif j == 1:
                cell.font = _DEPT_FONT
            else:
                cell.font = _BODY_FONT
            cell.alignment = _LEFT if j <= 1 else _CENTER
        name_cell = ws.cell(row=row, column=_SUM_FIRST)
        name_cell.hyperlink = Hyperlink(
            ref=name_cell.coordinate, location=f"'{sheet_name}'!A1"
        )
        name_cell.font = _LINK_FONT

    if report.overview_rows:
        first = get_column_letter(_SUM_FIRST)
        last_col = _SUM_FIRST + len(report.overview_columns) - 1
        last = get_column_letter(last_col)
        ws.auto_filter.ref = f"{first}{header_row}:{last}{row}"
        _outline(ws, header_row, _SUM_FIRST, row, last_col, inner_v=True)
    return row


_RANK_SLOTS = 3
# Each ranking column occupies a (name-col, value-col) pair on the B:I grid.
_RANK_SPANS = [(2, 3), (4, 5), (6, 7), (8, 9)]


def _summary_rankings(
    ws: Worksheet, report: AttendanceReport, *, start_row: int
) -> int:
    """Four podiums side by side, padded to three slots with em-dashes."""
    _section_header(ws, start_row, "TOP RANKINGS")
    ws.row_dimensions[start_row + 1].height = 6

    title_row = start_row + 2
    header_row = title_row + 1
    last_row = header_row + _RANK_SLOTS
    # The spacer row above the titles is filled too, so the section reads as
    # one continuous panel rather than a band with a white gap under it.
    for row in range(start_row + 1, last_row + 1):
        _paint(ws, row, _SUM_FIRST, _SUM_LAST, _ALT_FILL)

    for ranking, (name_col, value_col) in zip(report.rankings, _RANK_SPANS):
        ws.merge_cells(
            start_row=title_row,
            start_column=name_col,
            end_row=title_row,
            end_column=value_col,
        )
        title = ws.cell(row=title_row, column=name_col, value=ranking.title)
        title.font = _RANK_TITLE_FONT
        title.alignment = _LEFT

        col_head = ws.cell(row=header_row, column=name_col, value="Employee")
        col_head.font = _RANK_HEADER_FONT
        col_head.alignment = _LEFT
        val_head = ws.cell(row=header_row, column=value_col, value="Value")
        val_head.font = _RANK_HEADER_FONT
        val_head.alignment = _CENTER
        # A rule under the column heads is all that separates the podium.
        for col in (name_col, value_col):
            ws.cell(row=header_row, column=col).border = Border(bottom=_RULE_SIDE)

        for slot in range(_RANK_SLOTS):
            row = header_row + 1 + slot
            entry = ranking.entries[slot] if slot < len(ranking.entries) else None
            name_cell = ws.cell(
                row=row,
                column=name_col,
                value=f"{slot + 1}.  {entry.name if entry else '—'}",
            )
            name_cell.font = _RANK_NAME_FONT if entry else _RANK_EMPTY_FONT
            name_cell.alignment = _LEFT
            if entry is not None:
                value_cell = ws.cell(row=row, column=value_col, value=entry.value)
                value_cell.font = _RANK_VALUE_FONT
                value_cell.alignment = _CENTER
    return last_row


def _summary_footer(ws: Worksheet, row: int) -> None:
    _paint(ws, row, _SUM_FIRST, _SUM_LAST, _ALT_FILL)
    note = ws.cell(
        row=row,
        column=_SUM_FIRST,
        value="Generated from Prezlab attendance data.",
    )
    note.font = _FOOTER_FONT
    note.alignment = _LEFT
    brand = ws.cell(row=row, column=_SUM_LAST, value="prezlab:")
    brand.font = _FOOTER_BRAND_FONT
    brand.alignment = _RIGHT

# ---------- Per-employee sheets ----------

# Content sits on the same B:I grid as the summary, with a column A gutter.
_EMP_FIRST = 2  # B
_EMP_LAST = 9  # I
_EMP_WIDTHS = [2.44, 10.89, 11.55, 13, 11, 11, 11, 13, 8.89]

_DAY_COLUMNS = [
    "Date",
    "Day",
    "Status",
    "Check-in",
    "Check-out",
    "Worked",
    "Missing hours",
]

# The SUMMARY tile grid: three tiles per band at these column pairs.
_EMP_TILE_SPANS = [(2, 3), (4, 5), (6, 7)]


def _emp_section_header(ws: Worksheet, row: int, title: str) -> None:
    """A light band with a rule under it, spanning the sheet's content grid."""
    _paint(ws, row, _EMP_FIRST, _EMP_LAST, _ALT_FILL)
    for col in range(_EMP_FIRST, _EMP_LAST + 1):
        ws.cell(row=row, column=col).border = _SECTION_RULE
    cell = ws.cell(row=row, column=_EMP_FIRST, value=title)
    cell.font = _SECTION_FONT
    cell.alignment = _LEFT


def _render_employee_sheet(
    ws: Worksheet, section: EmployeeSection, period_label: str
) -> None:
    ws.sheet_view.showGridLines = False
    for idx, width in enumerate(_EMP_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _employee_header(ws, section, period_label)
    row = _employee_profile(ws, section, start_row=8)
    row = _employee_summary(ws, section, start_row=row + 2)
    row = _employee_daily_log(ws, section, start_row=row + 2)
    _employee_status_key(ws, start_row=row + 2)


def _employee_header(
    ws: Worksheet, section: EmployeeSection, period_label: str
) -> None:
    """Rows 2-5: dark band with the name, department, code and back-link."""
    for row in range(2, 6):
        _paint(ws, row, _EMP_FIRST, _EMP_LAST, _INK_FILL)
    ws.row_dimensions[2].height = 19.95
    ws.row_dimensions[3].height = 28.05
    ws.row_dimensions[4].height = 16.05

    brand = ws.cell(row=2, column=_EMP_FIRST, value="prezlab:")
    brand.font = Font(bold=True, size=11, color=_MINT)
    brand.alignment = _LEFT

    back = ws.cell(row=2, column=8, value="← Summary")
    back.hyperlink = Hyperlink(ref=back.coordinate, location=f"'{_SUMMARY_SHEET}'!A1")
    back.font = Font(bold=True, size=10, color=_MINT)
    back.alignment = _RIGHT

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    name = ws.cell(row=3, column=2, value=section.name)
    name.font = Font(bold=True, size=22, color="FFFFFF")
    name.alignment = _LEFT

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    caption = ws.cell(
        row=4,
        column=2,
        value=f"{section.department or DASH}  ·  Emp #{section.emp_code}",
    )
    caption.font = Font(size=11, color=_MINT)
    caption.alignment = _LEFT

    period_key = ws.cell(row=5, column=6, value="Period")
    period_key.font = _HERO_META_LABEL_FONT
    period_key.alignment = _RIGHT
    ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=8)
    period_value = ws.cell(row=5, column=7, value=period_label)
    period_value.font = Font(bold=True, size=10, color="FFFFFF")
    period_value.alignment = _RIGHT


def _employee_profile(
    ws: Worksheet, section: EmployeeSection, *, start_row: int
) -> int:
    """Identity on the left, working pattern on the right."""
    _emp_section_header(ws, start_row, "PROFILE")

    left = [
        ("Employee code", section.emp_code),
        ("Department", section.department or DASH),
        ("Pool / Team", section.pod or DASH),
    ]
    right = [
        ("Schedule", section.schedule_label),
        ("Shift start", section.shift_start),
        ("Shift end", section.shift_end),
    ]
    for i, ((llabel, lvalue), (rlabel, rvalue)) in enumerate(zip(left, right)):
        row = start_row + 1 + i
        _paint(ws, row, _EMP_FIRST, _EMP_LAST, _ALT_FILL)

        label = ws.cell(row=row, column=2, value=llabel)
        label.font = _PROFILE_LABEL_FONT
        label.alignment = _LEFT
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        value = ws.cell(row=row, column=3, value=lvalue)
        value.font = _PROFILE_VALUE_FONT
        value.alignment = _LEFT

        rlabel_cell = ws.cell(row=row, column=6, value=rlabel)
        rlabel_cell.font = _PROFILE_LABEL_FONT
        rlabel_cell.alignment = _LEFT
        rvalue_cell = ws.cell(row=row, column=7, value=rvalue)
        rvalue_cell.font = _PROFILE_VALUE_FONT
        rvalue_cell.alignment = _RIGHT
    return start_row + len(left)


def _employee_summary(
    ws: Worksheet, section: EmployeeSection, *, start_row: int
) -> int:
    """A grid of label / big-value / caption tiles, three per band."""
    _emp_section_header(ws, start_row, "SUMMARY")

    # (label, value, caption, highlighted) — the attendance tile is the one
    # the eye should land on, so it carries the mint fill.
    tiles = [
        (
            "Days attended",
            f"{section.days_attended} / {section.days_expected}",
            "attended of expected",
            False,
        ),
        (
            "Hours worked",
            section.total_worked,
            f"of {section.expected_worked} expected",
            False,
        ),
        ("Attendance", section.attendance_pct, "target 100%", True),
        (
            "Late arrivals",
            str(section.days_late),
            f"of {section.days_attended} present days",
            False,
        ),
        (
            "Avg worked / day",
            section.avg_worked,
            f"of {section.full_day} target",
            False,
        ),
        ("Days from home", str(section.days_wfh), "WFH days", False),
        (
            "Avg check-in",
            section.avg_arrival,
            f"shift starts {section.shift_start}",
            False,
        ),
        (
            "Avg check-out",
            section.avg_departure,
            f"shift ends {section.shift_end}",
            False,
        ),
        (
            "Missing sign-outs",
            str(section.missing_signouts),
            "no check-out logged",
            False,
        ),
        ("Absent", str(section.days_absent), "unplanned", False),
        ("On leave", str(section.days_on_leave), "approved", False),
        ("Public holidays", str(section.days_holiday), "in period", False),
    ]

    # A short tinted spacer above the grid, then three rows per tile band.
    _paint(ws, start_row + 1, _EMP_FIRST, _EMP_LAST, _ALT_FILL)
    ws.row_dimensions[start_row + 1].height = 4.05

    row = start_row + 2
    for band_start in range(0, len(tiles), 3):
        band = tiles[band_start : band_start + 3]
        label_row, value_row, caption_row = row, row + 1, row + 2
        ws.row_dimensions[label_row].height = 13.95
        ws.row_dimensions[value_row].height = 19.95
        ws.row_dimensions[caption_row].height = 13.05
        # The H:I margin stays tinted so the tiles read as cards on a panel.
        for band_row in (label_row, value_row, caption_row):
            _paint(ws, band_row, _EMP_FIRST, _EMP_LAST, _ALT_FILL)

        for (label, value, caption, highlight), (c0, c1) in zip(band, _EMP_TILE_SPANS):
            fill = _MINT_FILL if highlight else _WHITE_FILL
            for band_row in (label_row, value_row, caption_row):
                _paint(ws, band_row, c0, c1, fill)
            _tile_text(ws, label_row, c0, c1, label, _TILE_CAPTION_FONT)
            _tile_text(ws, value_row, c0, c1, value, _EMP_TILE_VALUE_FONT)
            _tile_text(ws, caption_row, c0, c1, caption, _TILE_CAPTION_FONT)
            # Box each tile so adjacent tiles read as separate cards.
            _outline(ws, label_row, c0, caption_row, c1)
        row = caption_row + 1
    return row - 1


def _tile_text(
    ws: Worksheet, row: int, c0: int, c1: int, value: str, font: Font
) -> None:
    ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
    cell = ws.cell(row=row, column=c0, value=value)
    cell.font = font
    cell.alignment = _LEFT


def _employee_daily_log(
    ws: Worksheet, section: EmployeeSection, *, start_row: int
) -> int:
    """Every day in the period, with a colored status pill."""
    _emp_section_header(ws, start_row, "DAILY LOG")

    row = start_row + 1
    header_row = row
    for i, name in enumerate(_DAY_COLUMNS):
        cell = ws.cell(row=row, column=_EMP_FIRST + i, value=name)
        cell.fill = _INK_FILL
        cell.font = _MINI_HEADER_FONT
        cell.alignment = _LEFT if i == 0 else _CENTER
    ws.row_dimensions[row].height = _ROW_HEIGHT
    last_col = _EMP_FIRST + len(_DAY_COLUMNS) - 1

    for i, line in enumerate(section.days):
        row += 1
        muted = line.status in _MUTED_STATUSES
        fill = _ALT_FILL if i % 2 else _WHITE_FILL
        values = [
            line.date,
            line.weekday,
            line.status,
            line.check_in,
            line.check_out,
            line.worked,
            line.missing_hours,
        ]
        for j, value in enumerate(values):
            cell = ws.cell(row=row, column=_EMP_FIRST + j, value=value)
            cell.fill = fill
            cell.font = _MUTED_FONT if muted else _BODY_FONT
            cell.alignment = _LEFT if j == 0 else _CENTER
        # The weekday reads as secondary even on a normal working day.
        ws.cell(row=row, column=3).font = _MUTED_FONT if muted else _DEPT_FONT
        _status_pill(ws.cell(row=row, column=4), line.status)

    # Rule the whole table, columns included, so the log stays readable.
    _outline(ws, header_row, _EMP_FIRST, row, last_col, inner_v=True)
    return row


def _status_pill(cell, status: str) -> None:
    """Paint one status cell as a filled pill, per the STATUS KEY palette."""
    style = _STATUS_STYLES.get(status)
    if style is None:
        return
    fill, color = style
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, size=9, color=color)
    cell.alignment = _CENTER


def _employee_status_key(ws: Worksheet, *, start_row: int) -> None:
    """The legend: one filled swatch per status, in the template's order."""
    title = ws.cell(row=start_row, column=_EMP_FIRST, value="STATUS KEY")
    _paint(ws, start_row, _EMP_FIRST, _EMP_LAST, _ALT_FILL)
    title.font = Font(bold=True, size=9, color=_INK)
    title.alignment = _LEFT

    key_row = start_row + 1
    _paint(ws, key_row, _EMP_FIRST, _EMP_LAST, _ALT_FILL)
    ws.row_dimensions[key_row].height = 24
    for i, status in enumerate(_STATUS_KEY):
        cell = ws.cell(row=key_row, column=_EMP_FIRST + i, value=status)
        _status_pill(cell, status)
        cell.font = Font(bold=True, size=8, color=_STATUS_STYLES[status][1])
        cell.border = Border(
            left=_RULE_SIDE, right=_RULE_SIDE, top=_RULE_SIDE, bottom=_RULE_SIDE
        )

    footer_row = key_row + 2
    _paint(ws, footer_row, _EMP_FIRST, _EMP_LAST, _ALT_FILL)
    note = ws.cell(
        row=footer_row,
        column=_EMP_FIRST,
        value="Generated from Prezlab attendance data.",
    )
    note.font = _FOOTER_FONT
    note.alignment = _LEFT
    brand = ws.cell(row=footer_row, column=_EMP_LAST, value="prezlab:")
    brand.font = _FOOTER_BRAND_FONT
    brand.alignment = _RIGHT
